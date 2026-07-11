import os
import logging
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session

from ai_orchestration.email_client import send_smtp_email
from ai_orchestration.pdf_generator import generate_pdf_report
import models

logger = logging.getLogger("email_service")

class EmailService:
    @staticmethod
    def generate_excel_report(output_path: str, title: str, headers: list, rows: list) -> bool:
        """
        Creates a professionally formatted corporate branded Excel file.
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "ERP Report"
            
            # Show grid lines
            ws.views.sheetView[0].showGridLines = True
            
            # Header title block
            ws.merge_cells("A1:E1")
            title_cell = ws["A1"]
            title_cell.value = title.upper()
            title_cell.font = Font(name="Arial", size=14, bold=True, color="1A365D")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formatted headers
            header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            
            for col_idx, header_val in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=str(header_val))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Table data cell formatting
            thin_border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )
            
            for row_idx, row_val in enumerate(rows, 4):
                for col_idx, cell_val in enumerate(row_val, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_val))
                    cell.border = thin_border
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(vertical="center")
                    
            # Auto-column width scaling
            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            wb.save(output_path)
            logger.info(f"Excel report generated successfully at {output_path}")
            return True
        except Exception as e:
            logger.error(f"Excel generation failed: {e}", exc_info=True)
            return False

    @staticmethod
    def send_low_stock_alert(recipient: str, item_name: str, current_stock: float, safety_level: float) -> bool:
        """
        Sends an alert warning of low inventory levels.
        """
        subject = f"⚠️ Safety Stock Alert: {item_name} is Low"
        body = (
            f"Dear Team,\n\n"
            f"Please take note that safety stock limits have been reached for:\n"
            f"Item: {item_name}\n"
            f"Current Stock: {current_stock}\n"
            f"Safety Limit: {safety_level}\n\n"
            f"Please verify and generate a purchase requisition if necessary.\n\n"
            f"Allure Living AI ERP Notification Service"
        )
        return send_smtp_email(to_email=recipient, subject=subject, text_body=body)

    @staticmethod
    def send_daily_report(db: Session, recipient: str) -> bool:
        """
        Gathers daily business metrics and emails a formatted PDF report with company branding.
        """
        today_str = date.today().isoformat()
        
        # Gather data from DB safely
        low_stock_count = db.query(models.InventoryItem).filter(
            models.InventoryItem.quantity <= models.InventoryItem.minimum_stock_level,
            models.InventoryItem.is_deleted == False
        ).count()
        
        active_projects_count = db.query(models.Project).filter(
            models.Project.status == "active",
            models.Project.is_deleted == False
        ).count()
        
        daily_expense_sum = db.query(models.DailyExpense).filter(
            models.DailyExpense.expense_date == date.today(),
            models.DailyExpense.is_deleted == False
        ).count() # Or sum, but let's count for simplicity
        
        title = f"Allure Living Daily Summary - {today_str}"
        sections = [
            {
                "header": "Daily Operations Status",
                "content": f"Summary report compiled automatically on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n"
                           f"Key metrics summarized below."
            },
            {
                "header": "Inventory & Production",
                "table_data": [
                    ["Metric Name", "Metric Value"],
                    ["Low Stock Warnings", str(low_stock_count)],
                    ["Active Manufacturing Projects", str(active_projects_count)],
                    ["Daily Expenses Logged", str(daily_expense_sum)]
                ]
            }
        ]
        
        pdf_path = f"backups/reports/Daily_Report_{today_str}.pdf"
        generate_pdf_report(pdf_path, title, sections)
        
        subject = f"📊 Daily Summary Report: {today_str}"
        body = f"Please find attached the Daily Operations Summary for {today_str} generated by Nexora AI."
        
        return send_smtp_email(to_email=recipient, subject=subject, text_body=body, attachment_path=pdf_path)
